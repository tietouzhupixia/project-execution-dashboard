import io

import pandas as pd
from openpyxl import load_workbook as load_xlsx

from src.personnel3_export import (
    ORANGE,
    VALUE_SHEETS,
    YELLOW,
    build_personnel3_formula_workbook,
    build_personnel3_value_workbook,
)
from src.personnel3_loader import Personnel3Inputs, coerce_number_series, load_personnel3_inputs
from src.personnel3_metrics import (
    MATCHED,
    NEEDS_CONFIRMATION,
    UNMATCHED,
    build_project_net_detail,
    match_outsource_projects,
)
from src.personnel3_outputs import (
    build_person_allocation,
    build_personnel3_checks,
    build_personnel3_exceptions,
    build_personnel3_list,
    build_personnel3_outputs,
    build_three_level_outputs,
    personnel3_department_people,
    personnel3_department_projects,
    personnel3_exceptions_of_type,
    personnel3_person_allocations,
    personnel3_project_allocations,
    personnel3_project_outsource,
    personnel3_project_rows,
)


def _implementation() -> pd.DataFrame:
    rows = [
        {
            "A-项目名称": "国网江苏电力2026年碳减排支撑服务",
            "项目管理编号": "P1",
            "A-项目经理": "甲",
            "A-项目经理区域": "华东事业部,电碳市场团队",
            "C-中标/合同金额": 1000,
            "预估项目金额": 900,
            "B-服务采购比例": 0.3,
            "开始执行日期": "2026-01-01",
            "预计验收日期（若已签约，默认经法）": "2026-12-31",
            "1/1进度": 0.1,
        },
        {
            "A-项目名称": "完全不同的西部研究项目",
            "项目管理编号": "P2",
            "A-项目经理": "乙",
            "A-项目经理区域": "数字化市场团队",
            "C-中标/合同金额": 0,
            "预估项目金额": 500,
            "B-服务采购比例": "20%",
            "开始执行日期": None,
            "预计验收日期（若已签约，默认经法）": None,
            "1/1进度": 0.2,
        },
    ]
    for row in rows:
        for i in range(1, 6):
            row[f"执行人员{i}"] = None
            row[f"执行人员{i}执行比例"] = None
    return pd.DataFrame(rows)


def _people() -> pd.DataFrame:
    return pd.DataFrame([{"人员 3": "张三", "所属区域 3": "华东事业部"}])


def test_load_three_input_sheets_by_name_without_virtual_ratio_fill():
    implementation = _implementation()
    outsource = pd.DataFrame([{"序号": 1, "外委项目名称": "外委A", "服务采购金额（元）": 100}])
    buffer = io.BytesIO()
    with pd.ExcelWriter(buffer, engine="openpyxl") as writer:
        implementation.to_excel(writer, sheet_name="input_实施进度表", index=False)
        outsource.to_excel(writer, sheet_name="input_外委更新金额", index=False)
        _people().to_excel(writer, sheet_name="input_人员关系表", index=False)
    buffer.seek(0)

    loaded = load_personnel3_inputs(buffer)

    assert loaded.ready
    assert loaded.source_sheets == {
        "implementation": "input_实施进度表",
        "outsource": "input_外委更新金额",
        "people": "input_人员关系表",
    }
    assert pd.isna(loaded.implementation.loc[0, "执行人员1执行比例"])
    assert "执行比例是否虚拟" not in loaded.implementation.columns


def test_missing_new_input_sheets_returns_business_errors():
    buffer = io.BytesIO()
    with pd.ExcelWriter(buffer, engine="openpyxl") as writer:
        _implementation().to_excel(writer, sheet_name="实施进度底表", index=False)
    buffer.seek(0)

    loaded = load_personnel3_inputs(buffer)

    assert not loaded.ready
    assert any("外委更新金额" in error for error in loaded.errors)
    assert any("人员关系表" in error for error in loaded.errors)


def test_matching_produces_three_states_and_accepts_only_matched():
    implementation = _implementation()
    outsource = pd.DataFrame(
        [
            {"序号": 1, "外委项目名称": "国网江苏电力2026年碳减排支撑服务", "服务采购金额（元）": 100},
            {"序号": 2, "外委项目名称": "国网江苏公司2026年碳减排支撑服务-13地市", "服务采购金额（元）": 200},
            {"序号": 3, "外委项目名称": "厨房装修采购", "服务采购金额（元）": 300},
        ]
    )

    result = match_outsource_projects(
        implementation,
        outsource,
        auto_threshold=0.99,
        confirmation_threshold=0.45,
    ).set_index("外委序号")

    assert result.loc["1", "匹配状态"] == MATCHED
    assert result.loc["1", "纳入采信金额"] == 100
    assert result.loc["2", "匹配状态"] == NEEDS_CONFIRMATION
    assert result.loc["2", "纳入采信金额"] == 0
    assert result.loc["3", "匹配状态"] == UNMATCHED
    assert result.loc["3", "纳入采信金额"] == 0


def test_manual_confirmation_changes_candidate_to_matched():
    outsource = pd.DataFrame(
        [{"序号": 2, "外委项目名称": "国网江苏公司2026年碳减排支撑服务-13地市", "服务采购金额（元）": 200}]
    )
    confirmations = pd.DataFrame(
        [{"外委序号": 2, "匹配状态": MATCHED, "对应实施项目编号": "P1"}]
    )

    result = match_outsource_projects(_implementation(), outsource, confirmations)

    assert result.loc[0, "匹配状态"] == MATCHED
    assert result.loc[0, "对应实施项目编号"] == "P1"
    assert result.loc[0, "纳入采信金额"] == 200
    assert "人工确认" in result.loc[0, "处理说明"]


def test_project_net_uses_matched_outsource_then_ratio_fallback_and_exclusion():
    matches = pd.DataFrame(
        [
            {
                "匹配状态": MATCHED,
                "对应实施项目编号": "P1",
                "纳入采信金额": 200,
            }
        ]
    )

    detail = build_project_net_detail(_implementation(), matches).set_index("项目管理编号")

    # P1: matched amount takes priority over 30%; net=800, annual delta=90%.
    assert detail.loc["P1", "净额取数基数"] == 1000
    assert detail.loc["P1", "最终采信服务采购金额"] == 200
    assert detail.loc["P1", "项目净执行合同额"] == 800
    assert detail.loc["P1", "26年净执行合同额"] == 720
    assert detail.loc["P1", "项目净额归属部门"] == "华东事业部"
    assert detail.loc["P1", "纳入口径26年净执行合同额"] == 720

    # P2: contract=0 -> estimate; no match -> ratio fallback; excluded region.
    assert detail.loc["P2", "净额取数基数"] == 500
    assert detail.loc["P2", "最终采信服务采购金额"] == 100
    assert detail.loc["P2", "26/12/31预计进度"] == 1
    assert detail.loc["P2", "26年净执行合同额"] == 320
    assert detail.loc["P2", "是否纳入口径"] == "否"
    assert detail.loc["P2", "纳入口径26年净执行合同额"] == 0


def test_cross_year_expected_progress_is_bounded_time_fraction():
    implementation = _implementation().iloc[[0]].copy()
    implementation.loc[implementation.index[0], "开始执行日期"] = "2026-01-01"
    implementation.loc[implementation.index[0], "预计验收日期（若已签约，默认经法）"] = "2027-01-01"
    implementation.loc[implementation.index[0], "1/1进度"] = 0
    empty_matches = pd.DataFrame(columns=["匹配状态", "对应实施项目编号", "纳入采信金额"])

    detail = build_project_net_detail(implementation, empty_matches)

    assert abs(detail.loc[0, "26/12/31预计进度"] - 364 / 365) < 1e-12
    assert 0 <= detail.loc[0, "26/12/31预计进度"] <= 1


def test_zero_base_never_turns_matched_outsource_into_negative_net():
    implementation = _implementation().iloc[[0]].copy()
    implementation.loc[implementation.index[0], "C-中标/合同金额"] = 0
    implementation.loc[implementation.index[0], "预估项目金额"] = None
    matches = pd.DataFrame(
        [{"匹配状态": MATCHED, "对应实施项目编号": "P1", "纳入采信金额": 550_000}]
    )

    detail = build_project_net_detail(implementation, matches)

    assert detail.loc[0, "净额取数基数"] == 0
    assert detail.loc[0, "外委子项目采信金额"] == 550_000
    assert detail.loc[0, "最终采信服务采购金额"] == 0
    assert detail.loc[0, "项目净执行合同额"] == 0
    assert detail.loc[0, "采信依据"] == "项目金额缺失，待补"


def test_money_parser_accepts_thousands_but_rejects_number_lists():
    parsed = coerce_number_series(pd.Series(["1,200,000", "600000,800000,900000", "20%"]), allow_percent=True)

    assert parsed.iloc[0] == 1_200_000
    assert pd.isna(parsed.iloc[1])
    assert parsed.iloc[2] == 0.2


def _phase2_fixture():
    implementation = _implementation()
    implementation.loc[0, "执行人员1"] = "张三"
    implementation.loc[0, "执行人员1执行比例"] = 0.6
    implementation.loc[0, "执行人员2"] = "范围外人员"
    implementation.loc[0, "执行人员2执行比例"] = 0.2
    implementation.loc[0, "执行人员3"] = "李四"
    implementation.loc[0, "执行人员3执行比例"] = None
    implementation.loc[1, "执行人员1"] = "王五"
    implementation.loc[1, "执行人员1执行比例"] = None
    people = pd.DataFrame(
        [
            {"人员 3": "张三", "所属区域 3": "华东事业部"},
            {"人员 3": "李四", "所属区域 3": "华东事业部"},
            {"人员 3": "王五", "所属区域 3": "西部业务部"},
            {"人员 3": "张三", "所属区域 3": "重复记录"},
        ]
    )
    matches = pd.DataFrame(
        [
            {
                "外委项目名称": "已确认外委",
                "匹配状态": MATCHED,
                "对应实施项目编号": "P1",
                "纳入采信金额": 200,
                "最高匹配度": 1.0,
            },
            {
                "外委项目名称": "待确认外委",
                "匹配状态": NEEDS_CONFIRMATION,
                "对应实施项目编号": "P2",
                "纳入采信金额": 0,
                "最高匹配度": 0.5,
            },
            {
                "外委项目名称": "未匹配外委",
                "匹配状态": UNMATCHED,
                "对应实施项目编号": "",
                "纳入采信金额": 0,
                "最高匹配度": 0.2,
            },
        ]
    )
    project_detail = build_project_net_detail(implementation, matches)
    return implementation, people, matches, project_detail


def test_personnel3_list_deduplicates_and_allocation_uses_only_filled_ratios():
    implementation, people, _, project_detail = _phase2_fixture()
    personnel3 = build_personnel3_list(people)
    allocation = build_person_allocation(implementation, project_detail, personnel3)

    assert personnel3["人员3"].tolist() == ["张三", "李四", "王五"]
    assert personnel3.set_index("人员3").loc["张三", "所属区域3"] == "华东事业部"
    assert allocation["执行人员"].tolist() == ["张三", "范围外人员"]
    assert allocation["是否人员3范围"].tolist() == ["是", "否"]
    assert allocation.loc[0, "分摊26年净执行合同额"] == 720 * 0.6
    assert "李四" not in allocation["执行人员"].tolist()
    assert "王五" not in allocation["执行人员"].tolist()


def test_three_level_outputs_keep_company_department_independent_of_person_gaps():
    implementation, people, matches, project_detail = _phase2_fixture()
    personnel3 = build_personnel3_list(people)
    allocation = build_person_allocation(implementation, project_detail, personnel3)
    company, department, person = build_three_level_outputs(project_detail, personnel3, allocation)

    # P2 is excluded, so company/department retain all of P1 despite only 80% ratios filled.
    assert company.loc[0, "26年净执行合同额"] == 720
    assert company.loc[0, "有效执行人数（人员3）"] == 3
    assert company.loc[0, "人均净合同额"] == 240
    assert company.loc[0, "项目数"] == 1
    east = department.set_index("部门").loc["华东事业部"]
    assert east["26年净执行合同额"] == 720
    assert east["有效执行人数（人员3）"] == 2
    assert east["人均净合同额"] == 360
    people_result = person.set_index("人员3")
    assert people_result.loc["张三", "26年净执行合同额"] == 432
    assert people_result.loc["李四", "参与分摊项目数"] == 0


def test_exceptions_cover_ratios_out_of_scope_matching_and_exclusion():
    implementation, people, matches, project_detail = _phase2_fixture()
    personnel3 = build_personnel3_list(people)
    allocation = build_person_allocation(implementation, project_detail, personnel3)
    exceptions = build_personnel3_exceptions(
        implementation, project_detail, personnel3, matches
    )
    kinds = exceptions["异常类型"].value_counts().to_dict()

    assert kinds["执行比例合计异常"] == 1
    assert kinds["执行人员未填分摊比例"] == 2
    assert kinds["缺少执行比例"] == 1
    assert kinds["分摊人员不在人员3范围"] == 1
    assert kinds["外委子项目需人工确认"] == 1
    assert kinds["外委子项目未匹配"] == 1
    assert kinds["不纳入口径项目"] == 1


def test_checks_and_combined_outputs_are_numeric_and_auditable():
    implementation, people, matches, project_detail = _phase2_fixture()
    outputs = build_personnel3_outputs(
        implementation,
        people,
        matches,
        project_detail,
        expected_people_count=3,
    )
    checks = outputs.checks.set_index("检查项")

    assert checks.loc["公司金额与部门合计一致", "差异/状态"] == 0
    assert checks.loc["人员3有效人数", "差异/状态"] == "通过"
    assert checks.loc["外委子项目自动采信金额", "计算值"] == 200
    assert abs(checks.loc["个人分摊覆盖率", "计算值"] - 0.8) < 1e-12
    assert outputs.summary.set_index("指标").loc["公司人均净合同额", "值"] == 240


def test_phase2_drilldown_helpers_share_output_scopes():
    implementation, people, matches, project_detail = _phase2_fixture()
    outputs = build_personnel3_outputs(
        implementation, people, matches, project_detail, expected_people_count=3
    )

    assert personnel3_department_projects(project_detail, "华东事业部")["项目管理编号"].tolist() == ["P1"]
    assert set(personnel3_department_people(outputs.personnel3_list, "华东事业部")["人员3"]) == {"张三", "李四"}
    assert personnel3_person_allocations(outputs.allocation, "张三")["项目管理编号"].tolist() == ["P1"]
    assert len(personnel3_project_rows(project_detail, "P1")) == 1
    assert len(personnel3_project_allocations(outputs.allocation, "P1")) == 2
    assert personnel3_project_outsource(matches, "P1")["外委项目名称"].tolist() == ["已确认外委"]
    assert len(personnel3_exceptions_of_type(outputs.exceptions, "执行比例合计异常")) == 1


def _export_fixture():
    implementation = _implementation()
    implementation["项目管理编号"] = implementation["项目管理编号"].astype(object)
    implementation.loc[0, "项目管理编号"] = 1001.0
    implementation.loc[0, "执行人员1"] = "张三"
    implementation.loc[0, "执行人员1执行比例"] = 0.6
    outsource = pd.DataFrame(
        [{"序号": 1, "外委项目名称": implementation.loc[0, "A-项目名称"], "服务采购金额（元）": 200}]
    )
    people = _people()
    inputs = Personnel3Inputs(implementation=implementation, outsource=outsource, people=people)
    matches = match_outsource_projects(implementation, outsource)
    detail = build_project_net_detail(implementation, matches)
    outputs = build_personnel3_outputs(implementation, people, matches, detail, expected_people_count=1)
    return inputs, matches, detail, outputs


def test_personnel3_value_export_contains_complete_audit_chain():
    inputs, matches, detail, outputs = _export_fixture()
    payload = build_personnel3_value_workbook(inputs, matches, detail, outputs)
    workbook = load_xlsx(io.BytesIO(payload), data_only=False)

    assert workbook.sheetnames == VALUE_SHEETS
    assert workbook["input_实施进度表"].max_row == len(inputs.implementation) + 1
    assert workbook["calculation_外委子项目匹配"].max_row == len(matches) + 1
    assert workbook["calculation_项目净额明细"].max_row == len(detail) + 1
    assert workbook["output_人员层面"].max_row == len(outputs.person) + 1
    assert workbook["output_公司层面"]["B2"].value == outputs.company.loc[0, "26年净执行合同额"]


def test_personnel3_formula_export_uses_live_formulas_styles_and_text_project_ids():
    inputs, matches, detail, outputs = _export_fixture()
    payload = build_personnel3_formula_workbook(inputs, matches, detail, outputs)
    workbook = load_xlsx(io.BytesIO(payload), data_only=False)

    project = workbook["calculation_项目净额明细"]
    allocation = workbook["calculation_人员分摊明细"]
    personnel3 = workbook["calculation_人员3名单"]
    match_sheet = workbook["calculation_外委子项目匹配"]
    company = workbook["output_公司层面"]
    person = workbook["output_人员层面"]
    checks = workbook["calculation_核验"]

    project_headers = {cell.value: cell.column for cell in project[1]}
    allocation_headers = {cell.value: cell.column for cell in allocation[1]}
    match_headers = {cell.value: cell.column for cell in match_sheet[1]}
    base_column = project_headers["净额取数基数"]
    allocation_amount_column = allocation_headers["项目26年净执行合同额"]

    assert workbook.sheetnames == VALUE_SHEETS
    assert project["A2"].value.startswith("=IF(AND('input_实施进度表'")
    assert '&""' in project["A2"].value
    assert allocation["A2"].value.startswith("=IF('input_实施进度表'")
    assert "'calculation_项目净额明细'!A2" in allocation["A2"].value
    assert personnel3["A2"].value.startswith("=IF(OR('input_人员关系表'")
    assert "COUNTIF" in personnel3["A2"].value
    assert "需人工确认" in match_sheet.cell(2, match_headers["匹配状态"]).value
    assert person["A2"].value.startswith("='calculation_人员3名单'")
    assert project.cell(2, base_column).value.startswith("=IF(")
    assert "SUMIF('calculation_项目净额明细'" in allocation.cell(2, allocation_amount_column).value
    assert company["B2"].value.startswith("=SUM(")
    assert checks["B2"].value.startswith("=")
    assert project.cell(1, base_column).fill.fgColor.rgb.endswith(ORANGE)
    assert project.cell(2, base_column).fill.fgColor.rgb.endswith(YELLOW)
    assert workbook.calculation.fullCalcOnLoad is True

    formula_required = [
        sheet for sheet in VALUE_SHEETS
        if sheet.startswith(("calculation_", "output_"))
    ]
    for sheet in formula_required:
        non_formulas = [
            cell.coordinate
            for row in workbook[sheet].iter_rows(min_row=2)
            for cell in row
            if not (isinstance(cell.value, str) and cell.value.startswith("="))
        ]
        assert not non_formulas, f"{sheet} contains non-formula data cells: {non_formulas[:5]}"

    # The formula export is a standalone offline template, not a fixed snapshot.
    assert workbook["input_实施进度表"].max_row == 201
    assert workbook["calculation_项目净额明细"].max_row == 201
    assert workbook["input_外委更新金额"].max_row == 201
    assert workbook["calculation_外委子项目匹配"].max_row == 201
    assert workbook["calculation_人员分摊明细"].max_row == 1001
    assert workbook["output_部门层面"].max_row == 101
    assert workbook["output_人员层面"].max_row == 101
    assert workbook["output_异常检查"].max_row == 3001
    assert workbook["calculation_项目净额明细"]["A201"].value.startswith("=IF(AND(")
    assert workbook["calculation_人员分摊明细"]["A1001"].value.startswith("=IF(")

    outsource_input = workbook["input_外委更新金额"]
    outsource_headers = [cell.value for cell in outsource_input[1]]
    for name in ["确认匹配状态", "确认对应实施项目编号", "导出时最高匹配度", "确认说明"]:
        assert name in outsource_headers
        column = outsource_headers.index(name) + 1
        assert outsource_input.cell(1, column).fill.fgColor.rgb.endswith(ORANGE)
        assert outsource_input.cell(2, column).fill.fgColor.rgb.endswith(YELLOW)
