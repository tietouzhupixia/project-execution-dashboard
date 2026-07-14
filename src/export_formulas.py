"""Audit-oriented Excel export: analysis cells are LIVE formulas that point back
to the raw sheet, so a human can inspect and re-verify the calculation logic.

Column letters are computed dynamically from the exported raw-sheet layout, so
the formulas stay correct even when the uploaded file reorders columns.
"""
from __future__ import annotations

import io
import math

import numpy as np
import pandas as pd
from openpyxl import Workbook
from openpyxl.utils import get_column_letter

RAW_SHEET = "实施进度底表"
RAW_REF = f"'{RAW_SHEET}'"
EFF_BASE_SHEET = "人效基础数据"
EFF_BASE_REF = f"'{EFF_BASE_SHEET}'"

# 归档动作列（派生）→ (进度阈值, 来源归档列)，公式在底表内重算
ACTION_FORMULA_RULES = [
    ("启动应归档", "启动已归档", 0.1, "启动归档"),
    ("中期应归档", "中期已归档", 0.5, "中期归档"),
    ("临近中期应归档", "临近中期已归档", 0.9, "临近终期归档"),
]

EXEC_PERSON_COLUMNS = [f"执行人员{i}" for i in range(1, 6)]
EXEC_RATIO_COLUMNS = [f"执行人员{i}执行比例" for i in range(1, 6)]


def build_formula_workbook(raw: pd.DataFrame, business_units: list[str], person: pd.DataFrame) -> bytes:
    """Assemble a workbook whose analysis sheets are live formulas over the raw sheet."""
    columns = list(raw.columns)
    n = len(raw)
    last_row = n + 1  # 表头占第 1 行，数据 2..n+1

    def letter(name: str) -> str | None:
        return get_column_letter(columns.index(name) + 1) if name in columns else None

    def rng(name: str) -> str | None:
        col = letter(name)
        return f"{RAW_REF}!${col}$2:${col}${last_row}" if col else None

    wb = Workbook()
    _write_raw_sheet(wb, raw, columns, last_row)
    _write_progress_sheet(wb, business_units, rng, last_row)
    _write_efficiency_base_sheet(wb, person, rng, last_row)
    _write_efficiency_sheet(wb, business_units, person)

    buffer = io.BytesIO()
    wb.save(buffer)
    return buffer.getvalue()


def _clean(value: object) -> object:
    """Make a pandas cell value safe for openpyxl."""
    if value is None:
        return None
    if isinstance(value, float) and math.isnan(value):
        return None
    try:
        if pd.isna(value):
            return None
    except (TypeError, ValueError):
        pass
    if isinstance(value, pd.Timestamp):
        return None if pd.isna(value) else value.to_pydatetime()
    if isinstance(value, np.integer):
        return int(value)
    if isinstance(value, np.floating):
        return float(value)
    if isinstance(value, np.bool_):
        return bool(value)
    return value


def _write_raw_sheet(wb: Workbook, raw: pd.DataFrame, columns: list[str], last_row: int) -> None:
    ws = wb.active
    ws.title = RAW_SHEET
    ws.append(columns)

    col_index = {name: i + 1 for i, name in enumerate(columns)}
    action_targets = {due for due, done, _, _ in ACTION_FORMULA_RULES} | {
        done for due, done, _, _ in ACTION_FORMULA_RULES
    }

    for r, (_, row) in enumerate(raw.iterrows(), start=2):
        for name in columns:
            if name in action_targets:
                continue  # 动作列用公式，稍后写
            ws.cell(row=r, column=col_index[name], value=_clean(row[name]))

    progress_col = get_column_letter(col_index["当前进度"]) if "当前进度" in col_index else None
    for due, done, threshold, source in ACTION_FORMULA_RULES:
        if due not in col_index:
            continue
        due_letter = get_column_letter(col_index[due])
        done_letter = get_column_letter(col_index[done])
        source_letter = get_column_letter(col_index[source]) if source in col_index else None
        for r in range(2, last_row + 1):
            if progress_col:
                ws[f"{due_letter}{r}"] = (
                    f"=IF(ISNUMBER({progress_col}{r}),IF({progress_col}{r}>={threshold},1,0),0)"
                )
            if source_letter:
                ws[f"{done_letter}{r}"] = (
                    f'=IF({due_letter}{r}=1,IF({source_letter}{r}="是",1,0),0)'
                )


def _write_progress_sheet(wb: Workbook, business_units: list[str], rng, last_row: int) -> None:
    ws = wb.create_sheet("进度信息分析")
    name_rng = rng("A-项目名称")
    region_rng = rng("A-项目经理区域")
    status_rng = rng("交付状态")
    prog_col = rng("当前进度")  # e.g. '实施进度底表'!$L$2:$L$N
    ac, ad, ae = rng("启动归档"), rng("中期归档"), rng("临近终期归档")

    def region_count(region_cell: str) -> str:
        return (
            f'=SUMPRODUCT(--ISNUMBER(SEARCH(","&{region_cell}&",",'
            f'","&SUBSTITUTE({region_rng},"，",",")&",")))'
        )

    def region_status_count(region_cell: str, keyword: str) -> str:
        return (
            f'=SUMPRODUCT(--ISNUMBER(SEARCH(","&{region_cell}&",",'
            f'","&SUBSTITUTE({region_rng},"，",",")&",")),'
            f'--ISNUMBER(SEARCH("{keyword}",{status_rng})))'
        )

    row = 1
    # —— 项目数 / 未验收 / 已验收 三块
    blocks = [
        ("一、项目数（按业务单元）", "项目数", None),
        ("二、未验收项目数（按业务单元）", "未验收项目数", "未验收"),
        ("三、已验收项目数（按业务单元）", "已验收项目数", "已验收"),
    ]
    for title, value_header, keyword in blocks:
        ws.cell(row=row, column=1, value=title)
        ws.cell(row=row + 1, column=1, value="业务单元")
        ws.cell(row=row + 1, column=2, value=value_header)
        ws.cell(row=row + 2, column=1, value="公司整体")
        if keyword is None:
            ws.cell(row=row + 2, column=2, value=f"=COUNTA({name_rng})" if name_rng else 0)
        else:
            ws.cell(
                row=row + 2, column=2,
                value=f'=SUMPRODUCT(--ISNUMBER(SEARCH("{keyword}",{status_rng})))' if status_rng else 0,
            )
        for i, unit in enumerate(business_units):
            rr = row + 3 + i
            ws.cell(row=rr, column=1, value=unit)
            cell = f"$A${rr}"
            ws.cell(row=rr, column=2, value=region_status_count(cell, keyword) if keyword else region_count(cell))
        row += 3 + len(business_units) + 1

    # —— 视角一：各阶段项目整体归档率
    ws.cell(row=row, column=1, value="四、视角一：各阶段项目整体归档率（当前及之前阶段都要完成）")
    row += 1
    headers = ["阶段", "进度范围", "应归档项目数（分母）", "已完成归档项目数（分子）", "归档率"]
    for c, h in enumerate(headers, start=1):
        ws.cell(row=row, column=c, value=h)
    v1_start = row + 1
    v1_rows = [
        ("启动阶段", "10%<=当前进度<50%", 0.1, 0.5, [ac]),
        ("中期阶段", "50%<=当前进度<90%", 0.5, 0.9, [ac, ad]),
        ("终期阶段", "当前进度>=90%", 0.9, None, [ac, ad, ae]),
    ]
    for i, (stage, scope, low, high, checks) in enumerate(v1_rows):
        rr = v1_start + i
        ws.cell(row=rr, column=1, value=stage)
        ws.cell(row=rr, column=2, value=scope)
        denom = f"--ISNUMBER({prog_col}),--({prog_col}>={low})"
        if high is not None:
            denom += f",--({prog_col}<{high})"
        ws.cell(row=rr, column=3, value=f"=SUMPRODUCT({denom})")
        num = denom + "".join(f',--({chk}="是")' for chk in checks)
        ws.cell(row=rr, column=4, value=f"=SUMPRODUCT({num})")
        ws.cell(row=rr, column=5, value=f"=IFERROR(D{rr}/C{rr},0)")
    total = v1_start + 3
    ws.cell(row=total, column=1, value="整体")
    ws.cell(row=total, column=2, value="当前进度>=10%")
    ws.cell(row=total, column=3, value=f"=SUM(C{v1_start}:C{v1_start + 2})")
    ws.cell(row=total, column=4, value=f"=SUM(D{v1_start}:D{v1_start + 2})")
    ws.cell(row=total, column=5, value=f"=IFERROR(D{total}/C{total},0)")
    row = total + 2

    # —— 视角二：环节维度归档完成率
    ws.cell(row=row, column=1, value="五、视角二：环节维度归档完成率（各归档环节单独计算）")
    row += 1
    headers2 = ["归档环节", "触发条件", "应完成归档环节数（分母）", "已完成归档环节数（分子）", "环节完成率"]
    for c, h in enumerate(headers2, start=1):
        ws.cell(row=row, column=c, value=h)
    v2_start = row + 1
    v2_rows = [
        ("启动归档环节", "当前进度>=10%", 0.1, ac),
        ("中期归档环节", "当前进度>=50%", 0.5, ad),
        ("终期归档环节", "当前进度>=90%", 0.9, ae),
    ]
    for i, (node, cond, threshold, archive_rng) in enumerate(v2_rows):
        rr = v2_start + i
        ws.cell(row=rr, column=1, value=node)
        ws.cell(row=rr, column=2, value=cond)
        denom = f"--ISNUMBER({prog_col}),--({prog_col}>={threshold})"
        ws.cell(row=rr, column=3, value=f"=SUMPRODUCT({denom})")
        ws.cell(row=rr, column=4, value=f'=SUMPRODUCT({denom},--({archive_rng}="是"))')
        ws.cell(row=rr, column=5, value=f"=IFERROR(D{rr}/C{rr},0)")
    total2 = v2_start + 3
    ws.cell(row=total2, column=1, value="整体")
    ws.cell(row=total2, column=2, value="三类环节合计")
    ws.cell(row=total2, column=3, value=f"=SUM(C{v2_start}:C{v2_start + 2})")
    ws.cell(row=total2, column=4, value=f"=SUM(D{v2_start}:D{v2_start + 2})")
    ws.cell(row=total2, column=5, value=f"=IFERROR(D{total2}/C{total2},0)")


def _write_efficiency_base_sheet(wb: Workbook, person: pd.DataFrame, rng, last_row: int) -> None:
    ws = wb.create_sheet(EFF_BASE_SHEET)
    ws.append(["人员", "净执行合同额", "所属区域/业务单元", "数据说明"])

    income = rng("收入")
    outsource = rng("B-服务采购比例")
    person_rngs = [rng(c) for c in EXEC_PERSON_COLUMNS]
    ratio_rngs = [rng(c) for c in EXEC_RATIO_COLUMNS]

    for i, (_, prow) in enumerate(person.iterrows(), start=2):
        ws.cell(row=i, column=1, value=str(prow["人员"]))
        if income and outsource and all(person_rngs) and all(ratio_rngs):
            # 不用 IFERROR(range,...)：SUMPRODUCT 里它会塌成标量而非逐元素，
            # 导致整列因子被当成单值。导出数据无错误值，空单元格在算术中即 0。
            terms = "+".join(
                f"(--({person_rngs[k]}=$A{i}))*{ratio_rngs[k]}" for k in range(5)
            )
            ws.cell(
                row=i, column=2,
                value=f"=SUMPRODUCT({income}*(1-{outsource})*({terms}))",
            )
        else:
            ws.cell(row=i, column=2, value=_clean(prow.get("净执行合同额")))
        ws.cell(row=i, column=3, value=str(prow.get("所属区域/业务单元", "")))
        ws.cell(row=i, column=4, value=str(prow.get("数据说明", "")))


def _write_efficiency_sheet(wb: Workbook, business_units: list[str], person: pd.DataFrame) -> None:
    ws = wb.create_sheet("人效分析")
    m = len(person) + 1  # 人效基础数据数据行 2..m
    net_rng = f"{EFF_BASE_REF}!$B$2:$B${m}"
    name_rng = f"{EFF_BASE_REF}!$A$2:$A${m}"
    area_rng = f"{EFF_BASE_REF}!$C$2:$C${m}"

    ws.cell(row=1, column=1, value="一、公司维度")
    company = [
        ("公司净执行合同额", f"=SUM({net_rng})"),
        ("统计人员数（全部人员）", f"=COUNTA({name_rng})"),
        ("有净额人数", f'=COUNTIF({net_rng},">0")'),
        ("人均净执行合同额（全部人员）", "=IFERROR(B2/B3,0)"),
        ("人均净执行合同额（有净额人员）", "=IFERROR(B2/B4,0)"),
    ]
    for i, (label, formula) in enumerate(company, start=2):
        ws.cell(row=i, column=1, value=label)
        ws.cell(row=i, column=2, value=formula)

    start = len(company) + 3
    ws.cell(row=start, column=1, value="二、业务单元维度")
    headers = ["业务单元", "净执行合同额", "人员数", "有净额人数", "人均净额（全部）", "人均净额（有净额）"]
    for c, h in enumerate(headers, start=1):
        ws.cell(row=start + 1, column=c, value=h)
    for i, unit in enumerate(business_units):
        rr = start + 2 + i
        cell = f"$A{rr}"
        ws.cell(row=rr, column=1, value=unit)
        ws.cell(row=rr, column=2, value=f"=SUMIF({area_rng},{cell},{net_rng})")
        ws.cell(row=rr, column=3, value=f"=COUNTIF({area_rng},{cell})")
        ws.cell(row=rr, column=4, value=f'=COUNTIFS({area_rng},{cell},{net_rng},">0")')
        ws.cell(row=rr, column=5, value=f"=IFERROR(B{rr}/C{rr},0)")
        ws.cell(row=rr, column=6, value=f"=IFERROR(B{rr}/D{rr},0)")
