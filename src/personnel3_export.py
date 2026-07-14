from __future__ import annotations

import io
import math

import numpy as np
import pandas as pd
from openpyxl import Workbook, load_workbook
from openpyxl.styles import Alignment, Font, PatternFill
from openpyxl.utils import get_column_letter

from .personnel3_loader import Personnel3Inputs
from .personnel3_outputs import Personnel3Outputs


BLUE = "1F4E78"
ORANGE = "C65911"
YELLOW = "FFF2CC"
WHITE = "FFFFFF"
MONEY_FORMAT = '#,##0.00'
PERCENT_FORMAT = '0.00%'

# The formula workbook is an offline template.  Users can fill these reserved
# input rows without returning to the web app; every downstream row already
# carries a formula.  Capacities are deliberately moderate to keep the file
# responsive in desktop Excel.
PROJECT_TEMPLATE_ROWS = 200
OUTSOURCE_TEMPLATE_ROWS = 200
PEOPLE_TEMPLATE_ROWS = 100
OUTSOURCE_CONFIRM_COLUMNS = [
    "确认匹配状态",
    "确认对应实施项目编号",
    "导出时最高匹配度",
    "确认说明",
]

VALUE_SHEETS = [
    "output_人均净合同额表",
    "output_公司层面",
    "output_部门层面",
    "output_人员层面",
    "output_异常检查",
    "calculation_外委子项目匹配",
    "calculation_项目净额明细",
    "calculation_人员分摊明细",
    "calculation_人员3名单",
    "calculation_核验",
    "input_实施进度表",
    "input_外委更新金额",
    "input_人员关系表",
]


def build_personnel3_value_workbook(
    inputs: Personnel3Inputs,
    matches: pd.DataFrame,
    project_detail: pd.DataFrame,
    outputs: Personnel3Outputs,
) -> bytes:
    """Export the current confirmed personnel-3 result as auditable values."""
    frames = _frames(inputs, matches, project_detail, outputs)
    buffer = io.BytesIO()
    with pd.ExcelWriter(buffer, engine="openpyxl") as writer:
        for sheet in VALUE_SHEETS:
            frames[sheet].to_excel(writer, sheet_name=sheet, index=False)
    workbook = load_workbook(io.BytesIO(buffer.getvalue()))
    for worksheet in workbook.worksheets:
        _style_sheet(worksheet, set())
    return _save(workbook)


def build_personnel3_formula_workbook(
    inputs: Personnel3Inputs,
    matches: pd.DataFrame,
    project_detail: pd.DataFrame,
    outputs: Personnel3Outputs,
) -> bytes:
    """Export a standalone Excel calculator driven only by the three input sheets."""
    frames = _formula_frames(inputs, matches, project_detail, outputs)
    workbook = Workbook()
    workbook.remove(workbook.active)
    formula_columns: dict[str, set[str]] = {}

    for sheet in VALUE_SHEETS:
        frame = frames[sheet]
        worksheet = workbook.create_sheet(sheet)
        _append_frame(worksheet, frame)

    formula_columns["calculation_外委子项目匹配"] = _write_match_formulas(
        workbook["calculation_外委子项目匹配"],
        workbook["input_外委更新金额"],
        workbook["calculation_项目净额明细"],
    )
    formula_columns["calculation_项目净额明细"] = _write_project_formulas(
        workbook["calculation_项目净额明细"],
        workbook["calculation_外委子项目匹配"],
        workbook["input_实施进度表"],
    )
    formula_columns["calculation_人员3名单"] = _write_personnel3_formulas(
        workbook["calculation_人员3名单"],
        workbook["input_人员关系表"],
    )
    formula_columns["calculation_人员分摊明细"] = _write_allocation_formulas(
        workbook["calculation_人员分摊明细"],
        workbook["calculation_项目净额明细"],
        workbook["calculation_人员3名单"],
        workbook["input_实施进度表"],
    )
    formula_columns["output_公司层面"] = _write_company_formulas(
        workbook["output_公司层面"],
        workbook["calculation_项目净额明细"],
        workbook["calculation_人员3名单"],
    )
    formula_columns["output_部门层面"] = _write_department_formulas(
        workbook["output_部门层面"],
        workbook["calculation_项目净额明细"],
        workbook["calculation_人员3名单"],
    )
    formula_columns["output_人员层面"] = _write_person_formulas(
        workbook["output_人员层面"],
        workbook["calculation_人员分摊明细"],
        workbook["calculation_人员3名单"],
    )
    formula_columns["output_人均净合同额表"] = _write_summary_formulas(
        workbook["output_人均净合同额表"], workbook["output_公司层面"]
    )
    formula_columns["calculation_核验"] = _write_check_formulas(
        workbook["calculation_核验"],
        workbook["output_公司层面"],
        workbook["output_部门层面"],
        workbook["calculation_人员分摊明细"],
        workbook["calculation_外委子项目匹配"],
    )
    formula_columns["output_异常检查"] = _write_exception_formulas(
        workbook["output_异常检查"],
        workbook["input_实施进度表"],
        workbook["calculation_项目净额明细"],
        workbook["calculation_人员3名单"],
        workbook["calculation_外委子项目匹配"],
    )

    for worksheet in workbook.worksheets:
        _style_sheet(worksheet, formula_columns.get(worksheet.title, set()))
    _style_confirmation_inputs(workbook["input_外委更新金额"])
    workbook.calculation.fullCalcOnLoad = True
    workbook.calculation.forceFullCalc = True
    workbook.calculation.calcMode = "auto"
    return _save(workbook)


def _frames(
    inputs: Personnel3Inputs,
    matches: pd.DataFrame,
    project_detail: pd.DataFrame,
    outputs: Personnel3Outputs,
) -> dict[str, pd.DataFrame]:
    return {
        "output_人均净合同额表": outputs.summary.copy(),
        "output_公司层面": outputs.company.copy(),
        "output_部门层面": outputs.department.copy(),
        "output_人员层面": outputs.person.copy(),
        "output_异常检查": outputs.exceptions.copy(),
        "calculation_外委子项目匹配": matches.copy(),
        "calculation_项目净额明细": project_detail.copy(),
        "calculation_人员分摊明细": outputs.allocation.copy(),
        "calculation_人员3名单": outputs.personnel3_list.copy(),
        "calculation_核验": outputs.checks.copy(),
        "input_实施进度表": inputs.implementation.copy(),
        "input_外委更新金额": inputs.outsource.copy(),
        "input_人员关系表": inputs.people.copy(),
    }


def _formula_frames(
    inputs: Personnel3Inputs,
    matches: pd.DataFrame,
    project_detail: pd.DataFrame,
    outputs: Personnel3Outputs,
) -> dict[str, pd.DataFrame]:
    """Create fixed-capacity row shells whose data cells are overwritten by formulas."""
    project_rows = max(PROJECT_TEMPLATE_ROWS, len(inputs.implementation))
    outsource_rows = max(OUTSOURCE_TEMPLATE_ROWS, len(inputs.outsource))
    people_rows = max(PEOPLE_TEMPLATE_ROWS, len(inputs.people))

    implementation = _pad_frame(inputs.implementation, project_rows)
    outsource = _pad_frame(inputs.outsource, outsource_rows)
    for name in OUTSOURCE_CONFIRM_COLUMNS:
        outsource[name] = pd.Series([""] * len(outsource), dtype=object)
    for index in range(min(len(matches), outsource_rows)):
        match = matches.iloc[index]
        outsource.at[index, "确认匹配状态"] = _clean(match.get("匹配状态"))
        outsource.at[index, "确认对应实施项目编号"] = _clean(
            match.get("对应实施项目编号")
        )
        outsource.at[index, "导出时最高匹配度"] = _clean(match.get("最高匹配度"))
        outsource.at[index, "确认说明"] = _clean(match.get("处理说明"))
    people = _pad_frame(inputs.people, people_rows)

    allocation_rows = project_rows * 5
    exception_rows = project_rows * 14 + outsource_rows
    return {
        "output_人均净合同额表": outputs.summary.copy(),
        "output_公司层面": outputs.company.copy(),
        "output_部门层面": _blank_like(outputs.department, people_rows),
        "output_人员层面": _blank_like(outputs.person, people_rows),
        "output_异常检查": _blank_like(outputs.exceptions, exception_rows),
        "calculation_外委子项目匹配": _blank_like(matches, outsource_rows),
        "calculation_项目净额明细": _blank_like(project_detail, project_rows),
        "calculation_人员分摊明细": _blank_like(outputs.allocation, allocation_rows),
        "calculation_人员3名单": _blank_like(outputs.personnel3_list, people_rows),
        "calculation_核验": outputs.checks.copy(),
        "input_实施进度表": implementation,
        "input_外委更新金额": outsource,
        "input_人员关系表": people,
    }


def _pad_frame(frame: pd.DataFrame, rows: int) -> pd.DataFrame:
    result = frame.reset_index(drop=True).copy()
    if len(result) < rows:
        result = pd.concat(
            [result, pd.DataFrame(index=range(rows - len(result)), columns=result.columns)],
            ignore_index=True,
        )
    return result


def _blank_like(frame: pd.DataFrame, rows: int) -> pd.DataFrame:
    return pd.DataFrame("", index=range(rows), columns=frame.columns)


def _append_frame(worksheet, frame: pd.DataFrame) -> None:
    worksheet.append(list(frame.columns))
    for row in frame.itertuples(index=False, name=None):
        worksheet.append([_clean(value) for value in row])


def _write_match_formulas(worksheet, input_sheet, project_sheet) -> set[str]:
    columns = _columns(worksheet)
    source = _columns(input_sheet)
    projects = _columns(project_sheet)
    project_last = max(project_sheet.max_row, 2)
    project_ids = _range(project_sheet.title, projects["项目管理编号"], 2, project_last)
    project_names = _range(project_sheet.title, projects["项目名称"], 2, project_last)
    for row in range(2, worksheet.max_row + 1):
        source_id = _source_ref(input_sheet, source.get("序号"), row)
        source_name = _source_ref(input_sheet, source["外委项目名称"], row)
        source_amount = _source_ref(input_sheet, source["服务采购金额（元）"], row)
        source_status = _source_ref(input_sheet, source["确认匹配状态"], row)
        source_project = _source_ref(
            input_sheet, source["确认对应实施项目编号"], row
        )
        source_score = _source_ref(input_sheet, source["导出时最高匹配度"], row)
        source_note = _source_ref(input_sheet, source["确认说明"], row)
        if source_id:
            worksheet[_cell(columns, "外委序号", row)] = (
                f'=IF({source_name}="","",IF({source_id}="","{row - 1}",{source_id}&""))'
            )
        else:
            worksheet[_cell(columns, "外委序号", row)] = (
                f'=IF({source_name}="","","{row - 1}")'
            )
        worksheet[_cell(columns, "外委项目名称", row)] = f'=IF({source_name}="","",{source_name})'
        worksheet[_cell(columns, "服务采购金额", row)] = (
            f'=IF({source_name}="","",IFERROR(1*{source_amount},0))'
        )
        status = _cell(columns, "匹配状态", row)
        project = _cell(columns, "对应实施项目编号", row)
        amount = _cell(columns, "服务采购金额", row)
        worksheet[status] = (
            f'=IF({source_name}="","",IF({source_status}="","需人工确认",{source_status}))'
        )
        worksheet[project] = f'=IF({source_name}="","",{source_project}&"")'
        worksheet[_cell(columns, "对应实施项目名称", row)] = (
            f'=IF({project}="","",IFERROR(INDEX({project_names},MATCH({project},{project_ids},0)),""))'
        )
        worksheet[_cell(columns, "最高匹配度", row)] = (
            f'=IF({source_name}="","",IFERROR(1*{source_score},0))'
        )
        worksheet[_cell(columns, "是否自动采信", row)] = (
            f'=IF({source_name}="","",AND({status}="已匹配",{project}<>""))'
        )
        worksheet[_cell(columns, "纳入采信金额", row)] = (
            f'=IF({source_name}="","",IF(AND({status}="已匹配",{project}<>""),{amount},0))'
        )
        worksheet[_cell(columns, "处理说明", row)] = (
            f'=IF({source_name}="","",IF({source_note}="",'
            f'"请在input_外委更新金额填写黄色确认列",{source_note}))'
        )
    return set(columns)


def _write_project_formulas(worksheet, match_sheet, input_sheet) -> set[str]:
    c = _columns(worksheet)
    m = _columns(match_sheet)
    source = _columns(input_sheet)
    match_last = max(match_sheet.max_row, 2)
    match_id = _range(match_sheet.title, m["对应实施项目编号"], 2, match_last)
    match_amount = _range(match_sheet.title, m["纳入采信金额"], 2, match_last)
    formula_columns = {
        "项目管理编号", "项目名称", "项目经理", "项目经理区域_原始",
        "项目净额归属部门", "中标/合同金额", "预计项目金额", "原服务采购比例",
        "开始执行日期", "预计验收日期", "1/1进度",
        "净额取数基数", "外委子项目采信金额", "最终采信服务采购金额", "采信依据",
        "项目净执行合同额", "26/12/31预计进度", "年度进度差", "26年净执行合同额",
        "是否纳入口径", "纳入口径26年净执行合同额",
    }
    for row in range(2, worksheet.max_row + 1):
        source_id = _source_ref(input_sheet, source["项目管理编号"], row)
        source_name = _source_ref(input_sheet, source["A-项目名称"], row)
        source_manager = _source_ref(input_sheet, source.get("A-项目经理"), row)
        source_region = _source_ref(input_sheet, source["A-项目经理区域"], row)
        source_contract = _source_ref(input_sheet, source["C-中标/合同金额"], row)
        source_estimate = _source_ref(input_sheet, source["预估项目金额"], row)
        source_ratio = _source_ref(input_sheet, source["B-服务采购比例"], row)
        source_start = _source_ref(input_sheet, source["开始执行日期"], row)
        source_end = _source_ref(
            input_sheet, source["预计验收日期（若已签约，默认经法）"], row
        )
        source_opening = _source_ref(input_sheet, source["1/1进度"], row)
        project_id = _cell(c, "项目管理编号", row)
        contract = _cell(c, "中标/合同金额", row)
        estimate = _cell(c, "预计项目金额", row)
        ratio = _cell(c, "原服务采购比例", row)
        start = _cell(c, "开始执行日期", row)
        end = _cell(c, "预计验收日期", row)
        opening = _cell(c, "1/1进度", row)
        base = _cell(c, "净额取数基数", row)
        matched = _cell(c, "外委子项目采信金额", row)
        accepted = _cell(c, "最终采信服务采购金额", row)
        net = _cell(c, "项目净执行合同额", row)
        expected = _cell(c, "26/12/31预计进度", row)
        delta = _cell(c, "年度进度差", row)
        annual = _cell(c, "26年净执行合同额", row)
        region = _cell(c, "项目经理区域_原始", row)
        included = _cell(c, "是否纳入口径", row)
        worksheet[project_id] = (
            f'=IF(AND({source_id}="",{source_name}=""),"",'
            f'IF({source_id}="","未填项目编号_行{row}",{source_id}&""))'
        )
        worksheet[_cell(c, "项目名称", row)] = f'={source_name}'
        worksheet[_cell(c, "项目经理", row)] = f'={source_manager}' if source_manager else '=""'
        worksheet[region] = f'={source_region}'
        worksheet[_cell(c, "项目净额归属部门", row)] = (
            f'=IF({project_id}="","",IF(AND(ISNUMBER(SEARCH("电碳市场团队",{region})),'
            f'SUBSTITUTE(SUBSTITUTE(SUBSTITUTE({region},"电碳市场团队",""),"，",","),",","")<>""),'
            f'TRIM(SUBSTITUTE(SUBSTITUTE(SUBSTITUTE({region},"电碳市场团队",""),"，",","),",","")),{region}))'
        )
        worksheet[contract] = f'=IF({project_id}="","",IFERROR(1*{source_contract},0))'
        worksheet[estimate] = f'=IF({project_id}="","",IFERROR(1*{source_estimate},0))'
        worksheet[ratio] = f'=IF({project_id}="","",IFERROR(1*{source_ratio},0))'
        worksheet[start] = f'=IF({project_id}="","",{source_start})'
        worksheet[end] = f'=IF({project_id}="","",{source_end})'
        worksheet[opening] = f'=IF({project_id}="","",IFERROR(1*{source_opening},0))'
        worksheet[base] = (
            f'=IF({project_id}="","",IF(OR({contract}="",{contract}=0),'
            f'IFERROR({estimate},0),{contract}))'
        )
        worksheet[matched] = f'=IF({project_id}="","",SUMIF({match_id},{project_id},{match_amount}))'
        worksheet[accepted] = (
            f'=IF({project_id}="","",IF({base}=0,0,IF({matched}>0,{matched},{ratio}*{base})))'
        )
        worksheet[_cell(c, "采信依据", row)] = (
            f'=IF({project_id}="","",IF({base}=0,"项目金额缺失，待补",'
            f'IF({matched}>0,"已匹配外委子项目合计","原服务采购比例×净额取数基数")))'
        )
        worksheet[net] = f'=IF({project_id}="","",{base}-{accepted})'
        worksheet[expected] = (
            f'=IF({project_id}="","",IF(OR({start}="",{end}=""),1,'
            f'IF({start}>DATE(2026,12,31),0,IF({end}<=DATE(2026,12,31),1,'
            f'MAX(0,MIN(1,(DATE(2026,12,31)-{start})/({end}-{start})))))))'
        )
        worksheet[delta] = f'=IF({project_id}="","",MAX(0,{expected}-{opening}))'
        worksheet[annual] = f'=IF({project_id}="","",{net}*{delta})'
        worksheet[included] = (
            f'=IF({project_id}="","",IF(OR(ISNUMBER(SEARCH("绿链",{region})),'
            f'ISNUMBER(SEARCH("数字化市场团队",{region}))),"否","是"))'
        )
        worksheet[_cell(c, "纳入口径26年净执行合同额", row)] = (
            f'=IF({project_id}="","",IF({included}="是",{annual},0))'
        )
    return formula_columns


def _write_personnel3_formulas(worksheet, input_sheet) -> set[str]:
    c = _columns(worksheet)
    source = _columns(input_sheet)
    for row in range(2, worksheet.max_row + 1):
        source_row = row
        source_person = _source_ref(input_sheet, source["人员 3"], source_row)
        source_department = _source_ref(input_sheet, source["所属区域 3"], source_row)
        person_letter = get_column_letter(source["人员 3"])
        seen_people = f"'{input_sheet.title}'!${person_letter}$2:{person_letter}{source_row}"
        worksheet[_cell(c, "人员3", row)] = (
            f'=IF(OR({source_person}="",COUNTIF({seen_people},{source_person})>1),"",{source_person})'
        )
        worksheet[_cell(c, "所属区域3", row)] = (
            f'=IF(OR({source_person}="",COUNTIF({seen_people},{source_person})>1),"",{source_department})'
        )
    return set(c)


def _write_allocation_formulas(
    worksheet,
    project_sheet,
    people_sheet,
    input_sheet,
) -> set[str]:
    c = _columns(worksheet)
    p = _columns(project_sheet)
    people = _columns(people_sheet)
    source = _columns(input_sheet)
    project_last = max(project_sheet.max_row, 2)
    people_last = max(people_sheet.max_row, 2)
    project_ids = _range(project_sheet.title, p["项目管理编号"], 2, project_last)
    project_names = _range(project_sheet.title, p["项目名称"], 2, project_last)
    project_amounts = _range(project_sheet.title, p["纳入口径26年净执行合同额"], 2, project_last)
    people_names = _range(people_sheet.title, people["人员3"], 2, people_last)
    people_departments = _range(people_sheet.title, people["所属区域3"], 2, people_last)
    sources = [
        (input_row, slot)
        for input_row in range(2, input_sheet.max_row + 1)
        for slot in range(1, 6)
    ]
    if len(sources) != max(worksheet.max_row - 1, 0):
        raise ValueError("人员分摊明细预留行数与实施进度表容量不一致。")
    for row, (input_row, slot) in enumerate(sources, start=2):
        source_person = _source_ref(input_sheet, source[f"执行人员{slot}"], input_row)
        source_ratio = _source_ref(
            input_sheet, source[f"执行人员{slot}执行比例"], input_row
        )
        project_id = _cell(c, "项目管理编号", row)
        person = _cell(c, "执行人员", row)
        amount = _cell(c, "项目26年净执行合同额", row)
        ratio = _cell(c, "执行比例", row)
        worksheet[project_id] = (
            f'=IF({source_person}="","",'
            f'{_sheet_ref(project_sheet.title, _cell(p, "项目管理编号", input_row))})'
        )
        worksheet[_cell(c, "项目名称", row)] = (
            f'=IF({project_id}="","",IFERROR(INDEX({project_names},MATCH({project_id},{project_ids},0)),""))'
        )
        worksheet[person] = f'=IF({source_person}="","",{source_person})'
        worksheet[ratio] = (
            f'=IF({person}="","",IF({source_ratio}="","",IFERROR(1*{source_ratio},"")))'
        )
        worksheet[_cell(c, "人员3部门", row)] = (
            f'=IF(OR({person}="",{ratio}=""),"",IFERROR(INDEX({people_departments},'
            f'MATCH({person},{people_names},0)),"人员3范围外"))'
        )
        worksheet[_cell(c, "是否人员3范围", row)] = (
            f'=IF(OR({person}="",{ratio}=""),"",'
            f'IF(COUNTIF({people_names},{person})>0,"是","否"))'
        )
        worksheet[amount] = (
            f'=IF(OR({project_id}="",{person}="",{ratio}=""),"",'
            f'SUMIF({project_ids},{project_id},{project_amounts}))'
        )
        worksheet[_cell(c, "分摊26年净执行合同额", row)] = (
            f'=IF({amount}="","",{amount}*{ratio})'
        )
    return set(c)


def _write_company_formulas(worksheet, project_sheet, people_sheet) -> set[str]:
    c, p = _columns(worksheet), _columns(project_sheet)
    project_last = max(project_sheet.max_row, 2)
    amounts = _range(project_sheet.title, p["纳入口径26年净执行合同额"], 2, project_last)
    included = _range(project_sheet.title, p["是否纳入口径"], 2, project_last)
    row = 2
    worksheet[_cell(c, "层级", row)] = '="公司"'
    amount = _cell(c, "26年净执行合同额", row)
    people = _cell(c, "有效执行人数（人员3）", row)
    worksheet[amount] = f'=SUM({amounts})'
    worksheet[people] = f'=COUNTIF(\'{people_sheet.title}\'!$A$2:$A${max(people_sheet.max_row, 2)},"?*")'
    worksheet[_cell(c, "人均净合同额", row)] = f'=IFERROR({amount}/{people},0)'
    worksheet[_cell(c, "项目数", row)] = f'=COUNTIF({included},"是")'
    return set(c)


def _write_department_formulas(worksheet, project_sheet, people_sheet) -> set[str]:
    c, p, pe = _columns(worksheet), _columns(project_sheet), _columns(people_sheet)
    project_last, people_last = max(project_sheet.max_row, 2), max(people_sheet.max_row, 2)
    departments = _range(project_sheet.title, p["项目净额归属部门"], 2, project_last)
    amounts = _range(project_sheet.title, p["纳入口径26年净执行合同额"], 2, project_last)
    included = _range(project_sheet.title, p["是否纳入口径"], 2, project_last)
    people_departments = _range(people_sheet.title, pe["所属区域3"], 2, people_last)
    for row in range(2, worksheet.max_row + 1):
        department = _cell(c, "部门", row)
        source_department = _sheet_ref(
            people_sheet.title, _cell(pe, "所属区域3", row)
        )
        department_letter = get_column_letter(pe["所属区域3"])
        seen_departments = (
            f"'{people_sheet.title}'!${department_letter}$2:{department_letter}{row}"
        )
        worksheet[department] = (
            f'=IF(OR({source_department}="",COUNTIF({seen_departments},'
            f'{source_department})>1),"",{source_department})'
        )
        amount = _cell(c, "26年净执行合同额", row)
        people = _cell(c, "有效执行人数（人员3）", row)
        worksheet[amount] = (
            f'=IF({department}="","",SUMIF({departments},{department},{amounts}))'
        )
        worksheet[people] = (
            f'=IF({department}="","",COUNTIF({people_departments},{department}))'
        )
        worksheet[_cell(c, "人均净合同额", row)] = (
            f'=IF({department}="","",IFERROR({amount}/{people},0))'
        )
        worksheet[_cell(c, "项目数", row)] = (
            f'=IF({department}="","",COUNTIFS({departments},{department},{included},"是"))'
        )
    return set(c)


def _write_person_formulas(worksheet, allocation_sheet, personnel3_sheet) -> set[str]:
    c, a, p = _columns(worksheet), _columns(allocation_sheet), _columns(personnel3_sheet)
    last = max(allocation_sheet.max_row, 2)
    people = _range(allocation_sheet.title, a["执行人员"], 2, last)
    scope = _range(allocation_sheet.title, a["是否人员3范围"], 2, last)
    amounts = _range(allocation_sheet.title, a["分摊26年净执行合同额"], 2, last)
    for row in range(2, worksheet.max_row + 1):
        person = _cell(c, "人员3", row)
        amount = _cell(c, "26年净执行合同额", row)
        count = _cell(c, "参与分摊项目数", row)
        worksheet[person] = f'={_sheet_ref(personnel3_sheet.title, _cell(p, "人员3", row))}'
        worksheet[_cell(c, "所属区域3", row)] = (
            f'={_sheet_ref(personnel3_sheet.title, _cell(p, "所属区域3", row))}'
        )
        worksheet[amount] = (
            f'=IF({person}="","",SUMIFS({amounts},{people},{person},{scope},"是"))'
        )
        worksheet[count] = (
            f'=IF({person}="","",COUNTIFS({people},{person},{scope},"是"))'
        )
        worksheet[_cell(c, "已填比例项目平均净额", row)] = (
            f'=IF({person}="","",IFERROR({amount}/{count},0))'
        )
        worksheet[_cell(c, "说明", row)] = (
            f'=IF({person}="","",IF({count}>0,"仅汇总已填执行比例",'
            f'"未填报分摊比例或未参与"))'
        )
    return set(c)


def _write_summary_formulas(worksheet, company_sheet) -> set[str]:
    c, company = _columns(worksheet), _columns(company_sheet)
    mapping = {
        "项目数": "项目数",
        "有效执行人数（人员3）": "有效执行人数（人员3）",
        "公司26年净执行合同额": "26年净执行合同额",
        "公司人均净合同额": "人均净合同额",
    }
    for row in range(2, worksheet.max_row + 1):
        metric = worksheet[_cell(c, "指标", row)].value
        worksheet[_cell(c, "指标", row)] = _constant_formula(metric)
        source = mapping.get(str(metric))
        if source:
            worksheet[_cell(c, "值", row)] = f"='{company_sheet.title}'!{_cell(company, source, 2)}"
        else:
            worksheet[_cell(c, "值", row)] = '=""'
    return set(c)


def _write_check_formulas(worksheet, company_sheet, department_sheet, allocation_sheet, match_sheet) -> set[str]:
    c = _columns(worksheet)
    company, department = _columns(company_sheet), _columns(department_sheet)
    allocation, match = _columns(allocation_sheet), _columns(match_sheet)
    company_amount = f"'{company_sheet.title}'!{_cell(company, '26年净执行合同额', 2)}"
    company_people = f"'{company_sheet.title}'!{_cell(company, '有效执行人数（人员3）', 2)}"
    department_amounts = _range(department_sheet.title, department["26年净执行合同额"], 2, max(department_sheet.max_row, 2))
    allocation_amounts = _range(allocation_sheet.title, allocation["分摊26年净执行合同额"], 2, max(allocation_sheet.max_row, 2))
    match_status = _range(match_sheet.title, match["匹配状态"], 2, max(match_sheet.max_row, 2))
    match_amounts = _range(match_sheet.title, match["纳入采信金额"], 2, max(match_sheet.max_row, 2))
    for row in range(2, worksheet.max_row + 1):
        name_cell = _cell(c, "检查项", row)
        note_cell = _cell(c, "说明", row)
        name = str(worksheet[name_cell].value)
        note = worksheet[note_cell].value
        worksheet[name_cell] = _constant_formula(name)
        worksheet[note_cell] = _constant_formula(note)
        calc = _cell(c, "计算值", row)
        compare = _cell(c, "对照值", row)
        status = _cell(c, "差异/状态", row)
        if name == "公司金额与部门合计一致":
            worksheet[calc] = f'={company_amount}'
            worksheet[compare] = f'=SUM({department_amounts})'
            worksheet[status] = f'={calc}-{compare}'
        elif name == "人员3有效人数":
            worksheet[calc] = f'={company_people}'
            worksheet[compare] = '=22'
            worksheet[status] = f'=IF({calc}={compare},"通过","需核对")'
        elif name == "外委子项目自动采信金额":
            worksheet[calc] = f'=SUMIF({match_status},"已匹配",{match_amounts})'
            worksheet[compare] = '=""'
            worksheet[status] = '="已匹配子项目合计"'
        elif name == "个人分摊覆盖率":
            worksheet[calc] = f'=IFERROR(SUM({allocation_amounts})/{company_amount},0)'
            worksheet[compare] = '=""'
            worksheet[status] = f'=IF(ABS({calc}-1)<0.000000001,"覆盖完整","未覆盖金额已在个人层面排除")'
    return set(c)


def _write_exception_formulas(
    worksheet,
    input_sheet,
    project_sheet,
    people_sheet,
    match_sheet,
) -> set[str]:
    """Build a fixed-capacity, formula-only exception report from the input rows."""
    c = _columns(worksheet)
    source = _columns(input_sheet)
    projects = _columns(project_sheet)
    people = _columns(people_sheet)
    matches = _columns(match_sheet)
    people_names = _range(
        people_sheet.title, people["人员3"], 2, max(people_sheet.max_row, 2)
    )
    output_row = 2

    def write_row(
        condition: str,
        project_id: str,
        name: str,
        kind_expression: str,
        note_expression: str,
    ) -> None:
        nonlocal output_row
        worksheet[_cell(c, "项目管理编号", output_row)] = (
            f'=IF({condition},{project_id},"")'
        )
        worksheet[_cell(c, "项目/外委子项目名称", output_row)] = (
            f'=IF({condition},{name},"")'
        )
        worksheet[_cell(c, "异常类型", output_row)] = (
            f'=IF({condition},{kind_expression},"")'
        )
        worksheet[_cell(c, "说明", output_row)] = (
            f'=IF({condition},{note_expression},"")'
        )
        output_row += 1

    for input_row in range(2, input_sheet.max_row + 1):
        project_id = _sheet_ref(
            project_sheet.title, _cell(projects, "项目管理编号", input_row)
        )
        project_name = _sheet_ref(
            project_sheet.title, _cell(projects, "项目名称", input_row)
        )
        included = _sheet_ref(
            project_sheet.title, _cell(projects, "是否纳入口径", input_row)
        )
        raw_project_id = _source_ref(
            input_sheet, source["项目管理编号"], input_row
        )
        raw_project_name = _source_ref(
            input_sheet, source["A-项目名称"], input_row
        )
        person_refs = [
            _source_ref(input_sheet, source[f"执行人员{slot}"], input_row)
            for slot in range(1, 6)
        ]
        ratio_refs = [
            _source_ref(input_sheet, source[f"执行人员{slot}执行比例"], input_row)
            for slot in range(1, 6)
        ]

        for person, ratio in zip(person_refs, ratio_refs):
            write_row(
                f'AND({person}<>"",{ratio}="")',
                project_id,
                project_name,
                _excel_string("执行人员未填分摊比例"),
                f'{person}&" 已填，但执行比例为空；未纳入个人分摊。"',
            )
            write_row(
                f'AND({person}<>"",COUNTIF({people_names},{person})=0)',
                project_id,
                project_name,
                _excel_string("分摊人员不在人员3范围"),
                f'{person}&" 不在人员关系表的人员 3 名单中；不进入个人层面。"',
            )

        persons = ",".join(person_refs)
        ratios = ",".join(ratio_refs)
        write_row(
            f'AND({raw_project_name}<>"",COUNTA({persons})>0,COUNT({ratios})=0)',
            project_id,
            project_name,
            _excel_string("缺少执行比例"),
            _excel_string("不使用 A-执行人员平均拆分；项目仅计入公司、部门项目净额。"),
        )
        write_row(
            f'AND({raw_project_name}<>"",COUNT({ratios})>0,ABS(SUM({ratios})-1)>0.000000001)',
            project_id,
            project_name,
            _excel_string("执行比例合计异常"),
            f'"已填写比例合计为 "&TEXT(SUM({ratios}),"0.00%")&"，个人层面按已填比例计算。"',
        )
        write_row(
            f'AND({raw_project_name}<>"",{raw_project_id}="")',
            project_id,
            project_name,
            _excel_string("缺少项目管理编号"),
            _excel_string("已按Excel行号生成临时项目编号。"),
        )
        write_row(
            f'{included}="否"',
            project_id,
            project_name,
            _excel_string("不纳入口径项目"),
            _excel_string(
                "项目经理区域包含绿链或数字化市场团队，不计入公司、部门、个人金额及项目数。"
            ),
        )

    for match_row in range(2, match_sheet.max_row + 1):
        status = _sheet_ref(match_sheet.title, _cell(matches, "匹配状态", match_row))
        project_id = _sheet_ref(
            match_sheet.title, _cell(matches, "对应实施项目编号", match_row)
        )
        name = _sheet_ref(
            match_sheet.title, _cell(matches, "外委项目名称", match_row)
        )
        score = _sheet_ref(
            match_sheet.title, _cell(matches, "最高匹配度", match_row)
        )
        condition = f'OR({status}="需人工确认",{status}="未匹配")'
        write_row(
            condition,
            project_id,
            name,
            f'IF({status}="需人工确认","外委子项目需人工确认","外委子项目未匹配")',
            f'"最高匹配度 "&TEXT({score},"0.0%")&"；本次不自动纳入最终采信服务采购金额。"',
        )

    if output_row != worksheet.max_row + 1:
        raise ValueError("异常检查预留行数与公式规则数量不一致。")
    return set(c)


def _excel_string(value: object) -> str:
    text = "" if value is None else str(value)
    return f'"{text.replace(chr(34), chr(34) * 2)}"'


def _constant_formula(value: object) -> str:
    if value is None or (isinstance(value, float) and math.isnan(value)):
        return '=""'
    if isinstance(value, (int, float, np.integer, np.floating)):
        return f'={value}'
    return f'={_excel_string(value)}'


def _style_confirmation_inputs(worksheet) -> None:
    headers = _columns(worksheet)
    for name in OUTSOURCE_CONFIRM_COLUMNS:
        column = headers[name]
        worksheet.cell(row=1, column=column).fill = PatternFill("solid", fgColor=ORANGE)
        worksheet.cell(row=1, column=column).font = Font(color=WHITE, bold=True)
        for row in range(2, worksheet.max_row + 1):
            worksheet.cell(row=row, column=column).fill = PatternFill("solid", fgColor=YELLOW)


def _style_sheet(worksheet, formula_columns: set[str]) -> None:
    headers = {cell.value: cell.column for cell in worksheet[1]}
    worksheet.freeze_panes = "A2"
    worksheet.auto_filter.ref = worksheet.dimensions
    for cell in worksheet[1]:
        color = ORANGE if cell.value in formula_columns else BLUE
        cell.fill = PatternFill("solid", fgColor=color)
        cell.font = Font(color=WHITE, bold=True)
        cell.alignment = Alignment(horizontal="center", vertical="center")
    for name in formula_columns:
        column = headers.get(name)
        if column:
            for row in range(2, worksheet.max_row + 1):
                worksheet.cell(row=row, column=column).fill = PatternFill("solid", fgColor=YELLOW)
    for column_cells in worksheet.columns:
        values = [str(cell.value) for cell in column_cells if cell.value is not None]
        header = str(column_cells[0].value or "")
        long_text = "名称" in header or "说明" in header or header == "采信依据"
        width_cap = 52 if long_text else 42
        width = min(max([len(value) for value in values] + [8]) + 2, width_cap)
        worksheet.column_dimensions[get_column_letter(column_cells[0].column)].width = width
        if long_text:
            for cell in column_cells[1:]:
                cell.alignment = Alignment(vertical="top", wrap_text=True)
                if cell.value is not None and len(str(cell.value)) > 30:
                    worksheet.row_dimensions[cell.row].height = max(
                        worksheet.row_dimensions[cell.row].height or 15,
                        30,
                    )
    worksheet.row_dimensions[1].height = 24
    for name, column in headers.items():
        number_format = _number_format(str(name))
        if number_format:
            for row in range(2, worksheet.max_row + 1):
                worksheet.cell(row=row, column=column).number_format = number_format


def _number_format(name: str) -> str | None:
    if "比例" in name or "进度" in name or "匹配度" in name or "覆盖率" in name:
        return PERCENT_FORMAT
    if "金额" in name or "净额" in name or name in {"计算值", "对照值", "差异/状态", "值"}:
        return MONEY_FORMAT
    return None


def _columns(worksheet) -> dict[str, int]:
    return {str(cell.value): cell.column for cell in worksheet[1] if cell.value is not None}


def _cell(columns: dict[str, int], name: str, row: int) -> str:
    return f"{get_column_letter(columns[name])}{row}"


def _range(sheet: str, column: int, first: int, last: int) -> str:
    letter = get_column_letter(column)
    return f"'{sheet}'!${letter}${first}:${letter}${last}"


def _sheet_ref(sheet: str, address: str) -> str:
    return f"'{sheet}'!{address}"


def _source_ref(worksheet, column: int | None, row: int) -> str | None:
    if column is None:
        return None
    return _sheet_ref(worksheet.title, f"{get_column_letter(column)}{row}")


def _clean(value: object) -> object:
    if value is None:
        return None
    if isinstance(value, float) and math.isnan(value):
        return None
    try:
        if pd.isna(value):
            return None
    except (TypeError, ValueError):
        pass
    if isinstance(value, pd.Timestamp):
        return value.to_pydatetime()
    if isinstance(value, np.integer):
        return int(value)
    if isinstance(value, np.floating):
        return float(value)
    if isinstance(value, np.bool_):
        return bool(value)
    return value


def _save(workbook: Workbook) -> bytes:
    buffer = io.BytesIO()
    workbook.save(buffer)
    return buffer.getvalue()
